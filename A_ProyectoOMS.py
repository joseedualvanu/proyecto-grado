"""
MODULO A
    Description: este es el módulo principal donde se ejecutan todos los modulos

    Args:
		--
    Returns:
        --
    Error:
        --
    Note:
        See https://www.datacamp.com/community/tutorials/docstrings-python
"""

from time import time
import datetime

# Obtengo la hora de comienzo para saber el tiempo de ejecución
start_time1 = time()
# Obtengo hora minutos y segundos para los nombres
start_time2 = datetime.datetime.now()
day = start_time2.day
month = start_time2.month
hour = start_time2.hour
year = start_time2.year
minute = start_time2.minute
second = start_time2.second
name_aux = str(day)+str(month)+str(year)+'_'+str(hour)+'\''+str(minute)+'\''+str(second)

# Parametros necesarios para las pruebas
fila = 250

# Numero de prueba
n = "S" + str(fila)

# Parametros necesarios para el modelo
# CB cantidad de bloques considerados
CB = 7
# Peso de los bloques anteriores a los elegidos en la BE_normalizada
Peso1 = 0.8
# Peso de los bloques posteriores a los elegidos en la BE_normalizada
Peso2 = 0.1
# Peso a divendo de bloques elegidos
Dividendo = 9
# Parametro que relaciona beneficio del cubo contra la clusterizacion por distancias en la FO
P = 0
# Capacidad de pedido por bloque
Capacidad = 50
# Capacidad = [0,30,30,30,50,70,70,70]
# Capacidad = [0,70,70,70,50,30,30,30]
# Capacidad = [0,80,80,50,50,20,20,20]
# Capacidad = [0,30,30,75,80,75,30,30]
# Cantidad de pedidos
# CantPedidosInput = int(CB * Capacidad * 0.8)
CantPedidosInput = 280
# Capacidad de armado de pedidos por bloque por persona
MO = 20
# Codigos postales definidos 15 actualmente
CP = [11000,11100,11200,11300,11400,11500,11600,11700,11800,11900,12000,12100,12300,15000,15800]

# Nombre del archivo con los pedidos y bloques
archivo_pedidos_1 = "0_Archivos/Pruebas/Pedido_BE_"+ n +".csv"
archivo_pedidos_2 = "0_Archivos/Pruebas/Pedido_PCP_" + n +".csv"
archivo_distancias = "0_Archivos/Pruebas/Distancias_CP_1.csv"
archivo_resultado = "0_Archivos/Pruebas/Resultado_" + n +".csv"
# Nombre de las graficas
Grafica1 = "0_Archivos/Pruebas/Grafica_Pedido_Bloque_" + n +".png"
Grafica2 = "0_Archivos/Pruebas/Grafica_Cumplimiento_" + n +".png"
Grafica3 = "0_Archivos/Pruebas/Grafica_CPBloque_" + n +".png"
Grafica4 = "0_Archivos/Pruebas/Grafica_BloquePersonal_" + n +".png"

"""
MODULO Auxiliar
    Se generan aleatoriamente los archivos necesarios
    Pedido_BE y Pedido_PCP
"""
from A_GeneradorPedidos import generadorpedidos
generadorpedidos(archivo_pedidos_1,archivo_pedidos_2,CantPedidosInput,CP,CB)

"""
MODULO B
    Se lee el archivo con pedidos y las elecciones
    Se crea la matriz BE pedido vs bloques elegidos
    Se crea la matriz TP pedido vs tipo
    Se crea la lista con los datos iniciales
"""
from B_ConvertorCSV_V2 import convertorcsv
(BE,TP,Lista_Datos_Iniciales,Lista_Distancias) = convertorcsv(CB,Dividendo,archivo_pedidos_1,archivo_distancias)

"""
MODULO C
    Se busca regularizar la matriz BE
    Sus filas sumen 10
"""
from C_Regularizador_cubo import regularizador_cubo
(CPBCP,Cant_Pedidos,Cant_Bloques,Cant_Codigos_Postales,Lista_Pedidos_CodPostales_Aux,Lista_Pedidos_CodPostales_Unicos,Lista_Pedidos_k,D , APuntaje,MPCP) = regularizador_cubo(BE, Peso1, Peso2, archivo_pedidos_2,Lista_Distancias)

"""
MODULO D
    Se ejecuta el modelo matemático
"""
from D_ModeloFinal import modelo
(X,A,ValorObjetivo)= modelo(CPBCP,Cant_Pedidos,Cant_Bloques,Cant_Codigos_Postales,TP,Lista_Pedidos_CodPostales_Aux,Lista_Pedidos_k,P,D,Capacidad)

"""
MODULO E
    Generacion de Dicts resultado
"""
from E_Resultado import resultado
(resultado, resultadoPorBloque, resultadoCPBloque , resultadoBloqueMO ,contTotal) = resultado(X , A , Cant_Pedidos,Cant_Bloques,Cant_Codigos_Postales, CPBCP,Lista_Pedidos_CodPostales_Aux , MO)

"""
MODULO F
    Se guarda el resultado en un csv
"""
from F_GuardarCSV import guardarcsv
guardarcsv(resultado,archivo_resultado)

"""
MODULO G
    Validación el resultado
"""
from G_ValidacionResultado import validacion
(resultadoValidacion,conteoValidacion2) = validacion(Lista_Datos_Iniciales,resultado)

# Guardo el tiempo transcurrido en el archivo con el resultado
file = open(archivo_resultado,"a")
file.write('Valor Objetivo: '+str(round(ValorObjetivo,2)) + '\n')
file.write('%Cumplimiento: '+str(conteoValidacion2['%Cumplimiento']) + '\n')
file.write('CP Totales: '+str(contTotal) +  '\n')

#Obtengo la hora de fin
elapsed_time = round(time() - start_time1,2)
print("Tiempo transcurrido: ",elapsed_time," segundos")

# from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(filename="resultado.xlsx")

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A' + str(fila)] = str(n)
ws['B' + str(fila)] = conteoValidacion2['%Cumplimiento']
ws['C' + str(fila)] = contTotal
ws['D' + str(fila)] = 'Ok'
ws['E' + str(fila)] = str(elapsed_time)

# Save the file
wb.save("resultado.xlsx")

"""
MODULO H
    Visualizar el resultado
"""
from H_Visualizacion_V2 import visualizacion
visualizacion(CB,resultado,resultadoValidacion,conteoValidacion2,resultadoCPBloque,n , resultadoBloqueMO, resultadoPorBloque, Grafica1, Grafica2, Grafica3, Grafica4)

#Imprimo numero de CP totales
print("CP Totales: " + str(contTotal))
print("P = " + str(P))
print("n = " + str(n))

file.write('Tiempo: '+str(elapsed_time)+' seg')
file.close()
